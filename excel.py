from csv import excel
from openpyxl import Workbook, load_workbook

# IMPORTANDO PLANILHA COM OS DADOS
def importa_planilha_excel(planilha):

    base = load_workbook(f"{planilha}")
    planilha = base.active
    return planilha

ws = Workbook()
etiqueta = ws.active



# FUNÇAO PARA RETORNAR o VALOR DA CELULA
def retorna_dado_celula(planilha, linha, coluna):
     celula  = planilha.cell(roll = linha, column = coluna)
     return celula.value


# FUNCAO PARA PREENCHER UMA CELULA
def preencher_celula(planilha,valor,linha, coluna):
    planilha.cell(row=linha, column= coluna, value = valor)


# FUNÇAO PARA RETORNAR UMA CELULA
def retorna_dado_celula(planilha, linha, coluna):
    celula  = planilha.cell(roll = linha, column = coluna)
    return celula.value

# FUNCAO PARA CRIAR UMA LISTA DE RE's
def retorna_lista_re(planilha):
    lista_re =list()
    for celula in planilha["A"]:
        if celula.value == "RE do usuário":
            continue
        else:
            lista_re.append(int(celula.value))

    return lista_re

# FUNÇAO PARA RETORNAR OS NOMES
def retorna_lista_nome(planilha):
    lista_nome = list()
    for celula in planilha["B"]:
        if celula.value == "Nome usuário":
            continue
        else:
            lista_nome.append(celula.value)
            
    return lista_nome

def gera_excel(siga , qtd_linha):
    for linha in range(qtd_linha -1):
        re = siga.funcionarios[linha].re
        nome = siga.funcionarios[linha].nome
        senha = siga.funcionarios[linha].senha
        data = siga.data
        numero = siga.numero

        if linha % 3 == 0:
            etiqueta.insert_rows(idx=linha, amount= 2 )
        else:    
            preencher_celula(etiqueta,re,linha,1)
            preencher_celula(etiqueta,nome,linha,2)
            preencher_celula(etiqueta,numero,linha,3)
            slinha = linha +1
            preencher_celula(etiqueta,senha,slinha,1)
            preencher_celula(etiqueta,data,slinha,2)

    ws.save('Etiqueta.xlsx')
        




