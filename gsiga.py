import excel
from banco_dados import conecta_bd
from gera_funcionario import gera_funcionario
from Siga import Siga
from openpyxl import load_workbook, Workbook

# IMPORTA PLANILHA EM EXCEL
planilha = excel.importa_planilha_excel("Base.xlsx");
etiqueta = Workbook()
etiqueta = etiqueta.create_sheet("Etiqueta")
# CONECTANDO NO BANCO DE DADOS

# conexao = conecta_bd()


# PROGRAMA PRINCIPAL --------------------------------------------------------------------------------------------------------------------
numero_siga = input("Nº SIGA: ")
operacao = input("OPERAÇÂO: ").upper()
data = input("DATA: ")

quantidade_linha = planilha.max_row
quantidade_coluna = planilha.max_column

lista_re = excel.retorna_lista_re(planilha)
lista_nome = excel.retorna_lista_nome(planilha)

# GERANDO UMA LISTA DE FUNCIONARIOS
lista_funcionarios = gera_funcionario(lista_re, lista_nome, quantidade_linha, operacao)

# GERANDO SIGA
siga = Siga(numero_siga, operacao, data, lista_funcionarios)

print(siga.funcionarios[0].re)

excel.gera_excel(siga, quantidade_linha)





















